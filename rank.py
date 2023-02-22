from selenium import webdriver
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.chrome.options import Options 
from bs4 import BeautifulSoup 
from openpyxl import Workbook 

# chrome driver (headless)
options = Options()
options.headless = True
driver = webdriver.Chrome(options=options)

# launch google and search keyword
driver.get("https://www.google.com")
search_box = driver.find_element("name", "q")
search_box.send_keys("seo")
search_box.send_keys(Keys.RETURN)

# scrape results 
soup = BeautifulSoup(driver.page_source, "html.parser")
results = soup.find_all("div", {"class": "g"})

# create excel workbook + sheet
workbook = Workbook()
worksheet = workbook.active

# headers for excel sheet 
worksheet.cell(row=1, column=1, value="Keyword")
worksheet.cell(row=1, column=2, value="Difficulty")

# difficulty estimate
for i, result in enumerate(results, start=2):
    title = result.find("h3").text
    url = result.find("a")["href"]

    keyword_count = title.lower().count("seo") + url.lower().count("seo")
    difficulty = round(keyword_count / len(results) * 100)

    # add to excel sheet
    worksheet.cell(row=i, column=1, value=title)
    worksheet.cell(row=i, column=2, value=difficulty)

workbook.save("keyword_difficulty.xlsx")

driver.quit()